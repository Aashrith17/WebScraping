import clointfusion as cf
import helium as h
from bs4 import BeautifulSoup
import openpyxl as op

path = r"C:\Users\Aashrith\OneDrive\Desktop\DevHack.xlsx"
 
wb_obj = op.load_workbook(path)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row

driver = cf.ChromeBrowser()
driver.open_browser()

for i in range(1,max_row+1):
    cell_obj_r = sheet_obj.cell(row = i, column = 1)
    driver.navigate(url="https://github.com/search?utf8=%E2%9C%93&q=&type=")
    cf.pause_program(seconds="2")
    driver.write(text=cell_obj_r.value,user_visible_text_element="Search Github")
    cf.pause_program(seconds="1")
    driver.hit_enter()
    h.set_driver(driver.browser_driver)
    element = h.find_all(h.S('#js-pjax-container > div > div.col-12.col-md-9.float-left.px-2.pt-3.pt-md-0.codesearch-results > div > div.d-flex.flex-column.flex-md-row.flex-justify-between.border-bottom.pb-3.position-relative > h3'))
    el = BeautifulSoup(str(element))
    [bool,res] = cf.string_extract_only_numbers(str(el.get_text()))
    cell_obj_w = sheet_obj.cell(row = i,column = 2)
    cell_obj_w.value = res
    wb_obj.save(path)
    cf.pause_program(seconds="2")